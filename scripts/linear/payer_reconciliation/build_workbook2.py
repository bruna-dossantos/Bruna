#!/usr/bin/env python3
import csv, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

WORK="/private/tmp/claude-501/-Users-brunadossantos-Projects-Bruna--claude-worktrees-inspiring-cannon-a54635/b4e7b3c8-2754-46fa-a3ad-6e0e1f6f10f1/scratchpad/recon"
detail=list(csv.DictReader(open(f"{WORK}/reconciliation2_detail.csv")))
summ=json.load(open(f"{WORK}/summary2.json"))
HDR=PatternFill("solid",fgColor="1F2937"); HF=Font(color="FFFFFF",bold=True); TITLE=Font(bold=True,size=15); SUB=Font(italic=True,size=9,color="6B7280")
COLORS={"DONE":"D1FAE5","FLIP":"FEF3C7","CONFLICT":"FEE2E2","NEW-TIX":"DBEAFE","UNIT-GAP":"F3E8FF"}
def hstyle(ws,n,row=1):
    for c in range(1,n+1): ws.cell(row,c).fill=HDR; ws.cell(row,c).font=HF
    ws.freeze_panes=ws.cell(row+1,1)
def widths(ws,ww):
    for i,w in enumerate(ww,1): ws.column_dimensions[get_column_letter(i)].width=w

wb=Workbook(); ws=wb.active; ws.title="Summary"
ws["A1"]="Order-Rule → Linear Reconciliation v2 (family-anchored)"; ws["A1"].font=TITLE
ws["A2"]="Source: Service Line Order Type Codes Jul 15 2026.csv · Linear cache Jul 8 · READ-ONLY, no Linear changes · 95.3% of rows resolved, 0 LOB mismatches"; ws["A2"].font=SUB
r=4
legend=[("DONE","Rule built & ticket already Done — in sync"),
        ("FLIP","Rule built, ticket Open → move to Done"),
        ("CONFLICT","Rule built, ticket Canceled/Not-Covered/Blocked → review"),
        ("NEW-TIX","Rule built, no ticket for this code/drug × payer → create"),
        ("UNIT-GAP","Code/drug not in Linear under that title → naming gap")]
for i,(k,d) in enumerate(legend):
    c=ws.cell(r+i,1,k); c.fill=PatternFill("solid",fgColor=COLORS[k]); c.font=Font(bold=True); ws.cell(r+i,2,d)
r=r+len(legend)+2
heads=["Verdict","DME units","DME rows","Infusion units","Infusion rows"]
for i,h in enumerate(heads,1): c=ws.cell(r,i,h); c.fill=HDR; c.font=HF
for i,v in enumerate(["DONE","FLIP","CONFLICT","NEW-TIX","UNIT-GAP"]):
    rr=r+1+i; c=ws.cell(rr,1,v); c.fill=PatternFill("solid",fgColor=COLORS[v]); c.font=Font(bold=True)
    ws.cell(rr,2,summ["buckets"].get(f"DME|{v}",0)); ws.cell(rr,3,summ["vol_buckets"].get(f"DME|{v}",0))
    ws.cell(rr,4,summ["buckets"].get(f"Infusion|{v}",0)); ws.cell(rr,5,summ["vol_buckets"].get(f"Infusion|{v}",0))
rr=r+7
ws.cell(rr,1,"Unresolved rows (no project mapping)").font=Font(italic=True)
ws.cell(rr,3,summ["unresolved"].get("DME",0)); ws.cell(rr,5,summ["unresolved"].get("Infusion",0))
widths(ws,[40,14,12,16,14])

def sheet(name,vs):
    ws=wb.create_sheet(name)
    heads=["Service","Code / Drug","Linear Project (payer)","Verdict","Rows built","Resolution basis","Current Linear state(s)","Ticket ID(s)","URL(s)"]
    ws.append(heads); hstyle(ws,len(heads))
    data=[d for d in detail if d["verdict"] in vs]; data.sort(key=lambda x:(x["service_line"],-int(x["rows_built"])))
    for d in data:
        ws.append([d["service_line"],d["unit"],d["project"],d["verdict"],int(d["rows_built"]),d["basis"],d["linear_states"],d["ticket_ids"],d["ticket_urls"]])
        ws.cell(ws.max_row,4).fill=PatternFill("solid",fgColor=COLORS.get(d["verdict"],"FFFFFF"))
    widths(ws,[9,26,40,10,10,26,26,16,52]); return len(data)

n1=sheet("FLIP → mark Done",{"FLIP"}); n2=sheet("NEW-TIX → create",{"NEW-TIX"})
n3=sheet("CONFLICT → review",{"CONFLICT"}); n4=sheet("UNIT-GAP → naming",{"UNIT-GAP"}); n5=sheet("DONE (in sync)",{"DONE"})
out=f"{WORK}/Order_Rule_Linear_Reconciliation_v3_Jul16.xlsx"; wb.save(out)
print("saved",out); print(f"FLIP={n1} NEW-TIX={n2} CONFLICT={n3} UNIT-GAP={n4} DONE={n5}")
