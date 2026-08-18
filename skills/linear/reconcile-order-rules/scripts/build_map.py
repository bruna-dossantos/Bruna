#!/usr/bin/env python3
"""Build the reproducible per-row "Order Type → Ticket" map (the shared Google Sheet artifact).

APPEND-ONLY schema: the raw export columns are preserved VERBATIM and in order (BASE_COLUMNS,
incl. the original v1_created_by_* creator columns — never dropped/reordered/renamed). New signals
are only APPENDED to the right (APPEND_COLUMNS). The publisher does clear-then-write, so only ever
ADD to APPEND_COLUMNS.

Called from reconcile.py, which already has `res` (raw + resolution) and the Linear ticket indexes.
PHASE-2 columns (ticket_is_parent/hierarchy/cycle/completed, duplicate_tickets) need a richer Linear
pull than the issues export provides — intentionally omitted, not fabricated.
"""
import csv

BASE_COLUMNS = [
    "code", "code_id", "order_rule_id", "order_rule_name", "existing_service_line",
    "mapped_service_lines", "mapping_status", "service_line_category", "plan_category",
    "payer_family", "insurance_payer", "eligibility_id", "plan_type", "payer_family_id",
    "plan_type_id", "insurance_payer_id", "Lookup_Key", "v1_order_rule_version_id",
    "v1_created_at", "v1_created_by_user_id", "v1_created_by_name", "v1_created_by_email",
]
APPEND_COLUMNS = [
    "hcpc", "criteria_creator_name", "criteria_creator_email", "state", "true_service_line",
    "resolved_project", "project_uuid", "resolution_basis",
    "verdict", "ticket_ids", "ticket_titles", "ticket_states", "ticket_urls",
]
COLUMNS = BASE_COLUMNS + APPEND_COLUMNS


def build(res, dme_idx, drug_idx, clean_code, norm, done_states, flippable, out_csv, out_xlsx, stamp):
    lin_codes = {c for c, _ in dme_idx}
    lin_drugs = {d for d, _ in drug_idx}

    def join(unit_raw, proj, cat):
        if not proj or cat not in ("DME", "Infusion"):
            return ("", "", "", "", "")
        is_dme = cat == "DME"
        idx = dme_idx if is_dme else drug_idx
        univ = lin_codes if is_dme else lin_drugs
        uk = clean_code(unit_raw) if is_dme else norm(unit_raw)
        if not uk:
            return ("", "", "", "", "")
        refs = idx.get((uk, proj), [])
        states = [a["state"] for a in refs]
        if uk not in univ:
            verdict = "UNIT-GAP"
        elif not states:
            verdict = "NEW-TIX"
        elif any(s in done_states for s in states):
            verdict = "DONE"
        elif any(s in flippable for s in states):
            verdict = "FLIP"
        else:
            verdict = "CONFLICT"
        return (verdict,
                ";".join(sorted({a["id"] for a in refs})),
                " | ".join(sorted({a.get("title", "") for a in refs if a.get("title")})),
                ";".join(sorted(set(states))),
                " ".join(sorted({a["url"] for a in refs})))

    out_rows = []
    for x in res:
        cat = x.get("service_line_category", "")
        unit = x.get("code", "") if cat == "DME" else x.get("true_service_line", "")
        verdict, ids, titles, sts, urls = join(unit, x.get("resolved_project", ""), cat)
        row = {c: x.get(c, "") for c in BASE_COLUMNS}   # raw export columns, preserved verbatim
        row.update({
            "hcpc": clean_code(x.get("code", "")),
            "criteria_creator_name": x.get("criteria_creator_name", ""),
            "criteria_creator_email": x.get("criteria_creator_email", ""),
            "state": x.get("state", ""),
            "true_service_line": x.get("true_service_line", ""),
            "resolved_project": x.get("resolved_project", ""),
            "project_uuid": x.get("project_uuid", ""),
            "resolution_basis": x.get("basis", ""),
            "verdict": verdict, "ticket_ids": ids, "ticket_titles": titles,
            "ticket_states": sts, "ticket_urls": urls,
        })
        out_rows.append(row)

    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        w.writeheader()
        w.writerows(out_rows)

    if out_xlsx:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            import re as _re
            ILLEGAL = _re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
            wb = Workbook(); ws = wb.active; ws.title = "Order Type → Ticket"
            ws.append(COLUMNS)
            hdr = PatternFill("solid", fgColor="1F2937"); hf = Font(color="FFFFFF", bold=True)
            for c in range(1, len(COLUMNS) + 1):
                ws.cell(1, c).fill = hdr; ws.cell(1, c).font = hf
            ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
            for r in out_rows:
                ws.append([ILLEGAL.sub('', str(r[c] or '')) for c in COLUMNS])
            wb.save(out_xlsx)
        except Exception as e:
            print(f"  (build_map: xlsx skipped — {e})")

    return out_rows
