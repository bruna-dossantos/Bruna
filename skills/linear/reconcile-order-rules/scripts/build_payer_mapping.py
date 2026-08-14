#!/usr/bin/env python3
"""Render the unique payer -> Linear project mapping with a 0-5 confidence score.

Standard output of the reconcile-order-rules skill. Dedupes the resolved rows on the
mapping-determining key (payer_family x insurance_payer x plan_category) and assigns a
confidence derived from the resolution layer that won for that payer:

  5  crosswalk                  — Bruna's hand-validated source of truth
  4  family+cat / +state        — deterministic family/category(+state) rule
  3  ordername / bcbs->anthem    — order-name mined brand+state, or BCBS-blank fallback
  3  name:mapper (mapper conf 3) — payer-linear-mapper strong match
  2  name:mapper (mapper conf 2) / name:norm — weaker name-based match
  0  unresolved                 — needs review

The mapper's own numeric confidence is read back from the .recon_work/*_mapping.csv files
(when present) so name:mapper rows keep the mapper's distinction between strong and weak.
"""
import csv, os, sys
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import resolver as R

HDR=PatternFill("solid",fgColor="1F2937"); HF=Font(color="FFFFFF",bold=True)
CONF_FILL={5:"D1FAE5",4:"D1FAE5",3:"FEF3C7",2:"FED7AA",0:"FEE2E2"}
# `validated` is the feedback round-trip column (Bruna marks x/yes and re-feeds via apply_feedback.py).
# `order_type_names` is LAST and only populated when the order name actually drove the mapping.
COLS=["payer_family","insurance_payer","plan_category","service_line_category",
      "resolved_project","project_uuid","confidence","confidence_label","basis","volume_rows",
      "validated","order_type_names"]

def _order_contributed(raw_basis, payer, orn):
    """True only when the order_rule_name is what drove this row's project resolution."""
    if raw_basis=="ordername": return True
    # family+state / family+cat+state: the order name contributed only if the state came from it
    # (resolve_family reads state from the payer first, then the order name).
    if raw_basis in ("family+state","family+cat+state"):
        return R.parse_state(payer) is None and R.parse_state(orn) is not None
    return False

def _load_mapper_conf(workdir):
    """source payer name -> mapper numeric confidence (from the payer-linear-mapper output)."""
    conf={}
    import os
    for fn in ["dme_mapping.csv","inf_mapping.csv"]:
        p=f"{workdir}/{fn}"
        if not os.path.exists(p): continue
        for r in csv.DictReader(open(p)):
            try: conf[r.get('source','').strip()]=int(r.get('confidence',0))
            except: pass
    return conf

def _score(raw_basis, payer, mapconf):
    b=raw_basis
    if b=="crosswalk": return 5,"Very high — hand-validated crosswalk"
    if b in ("family+cat","family+cat+state","family+state"): return 4,"High — family/category(+state) rule"
    if b=="ordername": return 3,"Medium — order-name mined brand+state"
    if b=="family:bcbs-blank->anthem": return 3,"Medium — BCBS blank→Anthem fallback"
    if b=="name:mapper":
        c=mapconf.get(payer,2)
        return (3,"Medium — name mapper (strong)") if c>=3 else (2,"Low-medium — name mapper")
    if b=="name:norm": return 2,"Low-medium — exact/unique name normalize"
    return 0,"Unresolved — needs review"

def build(res, workdir, csv_path, xlsx_path):
    """res: list of resolved row dicts from reconcile.main(). Writes CSV + XLSX; returns row count."""
    mapconf=_load_mapper_conf(workdir)
    # Key on (family × payer × category × resolved_project) so each row is internally consistent:
    # if one payer combo splits across projects (or partly resolves), each project — and the
    # unresolved remainder — is its own row, with matching basis/confidence/order names.
    agg=defaultdict(lambda:{"vol":0,"ordernames":set()})
    for r in res:
        proj=(r.get("resolved_project") or "").strip()
        key=(r['payer_family'].strip(), r['insurance_payer'].strip(), r['plan_category'].strip(), proj)
        a=agg[key]; a["vol"]+=1
        a["service_line_category"]=r['service_line_category'].strip()
        a["project_uuid"]=(r.get("project_uuid") or "").strip()
        a["raw_basis"]=r.get("basis") or ""
        a["basis"]=a["raw_basis"].replace("unresolved:","")
        orn=(r.get("order_rule_name") or "").strip()
        if orn and _order_contributed(a["raw_basis"], key[1], orn):
            a["ordernames"].add(orn)

    out=[]
    for (fam,pay,cat,proj),a in agg.items():
        sc,lab=_score(a["raw_basis"], pay, mapconf)
        out.append({"payer_family":fam,"insurance_payer":pay,"plan_category":cat,
            "service_line_category":a["service_line_category"],"resolved_project":proj,
            "project_uuid":a["project_uuid"],"confidence":sc,"confidence_label":lab,
            "basis":a["basis"],"volume_rows":a["vol"],"validated":"",
            "order_type_names":"; ".join(sorted(a["ordernames"]))})
    out.sort(key=lambda x:(-x["confidence"], -x["volume_rows"], x["insurance_payer"]))

    with open(csv_path,"w",newline="") as f:
        wt=csv.DictWriter(f,fieldnames=COLS,extrasaction="ignore"); wt.writeheader(); wt.writerows(out)

    wb=Workbook(); ws=wb.active; ws.title="Payer → Project"
    heads=["Payer Family","Insurance Payer","Plan Category","Service Line","Linear Project",
           "Project UUID","Confidence (0-5)","Confidence Label","Basis","Volume (rows)",
           "Validated (x)","Order Type Name(s) — if it drove the mapping"]
    ws.append(heads)
    for c in range(1,len(heads)+1): ws.cell(1,c).fill=HDR; ws.cell(1,c).font=HF
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(heads))}1"
    for r in out:
        ws.append([r[c] for c in COLS])
        ws.cell(ws.max_row,7).fill=PatternFill("solid",fgColor=CONF_FILL.get(r["confidence"],"FFFFFF"))
    for i,wd in enumerate([22,34,16,12,42,38,14,34,26,12,12,60],1):
        ws.column_dimensions[get_column_letter(i)].width=wd
    wb.save(xlsx_path)

    dist=Counter(r["confidence"] for r in out)
    return out, dist
