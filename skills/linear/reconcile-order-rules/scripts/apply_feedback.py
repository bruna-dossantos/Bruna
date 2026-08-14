#!/usr/bin/env python3
"""Feed a validated payer→project mapping back into the crosswalk (source of truth).

Reads an edited `Payer_Project_Mapping_<date>.(xlsx|csv)`, takes every row Bruna marked in the
`validated` column, and appends/updates the matching entry in `payer_project_crosswalk.csv`.
Those entries then win automatically on the next reconcile run.

The `validated` cell is interpreted three ways:
  • truthy word (x / y / yes / true / 1 / ✓) — validate the row as-is: project = `resolved_project`
    (corrected in place if Bruna edited it), UUID from `project_uuid` else looked up by name.
  • a Linear project UUID — AUTHORITATIVE mapping. The project name is taken from that UUID via
    insurance_projects.csv (overriding whatever is in the row's project column), so this both
    resolves a blank row and corrects a wrong auto-mapping.
  • "New Project Needed" (any text containing "new project") — there is no existing Linear project
    to point at; NOT written to the crosswalk. Collected into a needs-new-project report instead.

Backs up the crosswalk before writing. Idempotent: re-feeding the same file is a no-op.

Usage:
  python3 apply_feedback.py "<validated mapping>.(xlsx|csv)" [--dry-run]
  python3 apply_feedback.py "<validated mapping>" --crosswalk <path> --projects <path>
"""
import csv, os, sys, re, shutil, datetime, argparse

HOME=os.path.expanduser("~")
MASTER=f"{HOME}/Documents/Claude/Projects/Linear Master Data"
OUT=f"{HOME}/Documents/Claude/Projects/Criteria Updates"
CROSSWALK=f"{MASTER}/payer_project_crosswalk.csv"
PROJECTS=f"{MASTER}/insurance_projects.csv"
XCOLS=["payer_family","insurance_payer","plan_category","linear_project","project_uuid","source","effective"]
TRUTHY={"x","y","yes","true","1","✓","validated","done"}
UUID_RE=re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', re.I)
# map the workbook's display headers (and the machine csv headers) onto canonical field names
ALIASES={
 "payer family":"payer_family","payer_family":"payer_family",
 "insurance payer":"insurance_payer","insurance_payer":"insurance_payer",
 "plan category":"plan_category","plan_category":"plan_category",
 "linear project":"resolved_project","resolved_project":"resolved_project",
 "project uuid":"project_uuid","project_uuid":"project_uuid",
 "validated (x)":"validated","validated":"validated",
}

def read_rows(path):
    """Yield canonical-keyed dicts from an .xlsx or .csv feedback file."""
    if path.lower().endswith((".xlsx",".xlsm")):
        from openpyxl import load_workbook
        ws=load_workbook(path, data_only=True).active
        it=ws.iter_rows(values_only=True)
        hdr=[ALIASES.get(str(h).strip().lower(), str(h).strip().lower()) if h is not None else "" for h in next(it)]
        for raw in it:
            yield {hdr[i]:("" if v is None else str(v).strip()) for i,v in enumerate(raw) if i<len(hdr)}
    else:
        for r in csv.DictReader(open(path)):
            yield {ALIASES.get((k or '').strip().lower(),(k or '').strip().lower()):(v or '').strip() for k,v in r.items()}

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("feedback"); ap.add_argument("--crosswalk",default=CROSSWALK)
    ap.add_argument("--projects",default=PROJECTS); ap.add_argument("--dry-run",action="store_true")
    a=ap.parse_args()
    stamp=datetime.date.today().isoformat(); month=stamp[:7]

    proj_by_name={r['Name'].strip():r['UUID'].strip() for r in csv.DictReader(open(a.projects))}
    proj_by_uuid={v:k for k,v in proj_by_name.items()}
    # normalized index (case/punctuation/&-vs-and insensitive) for tolerant, exact-normalized lookup
    def pnorm(s): return re.sub(r'[^a-z0-9]+','',re.sub(r'\band\b',' ',(s or '').lower().replace('&',' and ')))
    proj_by_norm={}
    for n,u in proj_by_name.items(): proj_by_norm.setdefault(pnorm(n),[]).append((n,u))

    xrows=[]; xindex={}
    if os.path.exists(a.crosswalk):
        for r in csv.DictReader(open(a.crosswalk)):
            xrows.append(r)
            xindex[(r['payer_family'].strip(),r['insurance_payer'].strip(),r['plan_category'].strip())]=r
    else:
        print(f"WARNING: crosswalk not found at {a.crosswalk} — creating a new one.")

    added=updated=skipped_unmarked=skipped_same=0; problems=[]; new_projects=[]
    for r in read_rows(a.feedback):
        v=(r.get("validated") or "").strip()
        if not v:
            skipped_unmarked+=1; continue
        fam=(r.get("payer_family") or "").strip()
        pay=(r.get("insurance_payer") or "").strip()
        cat=(r.get("plan_category") or "").strip()
        who=f"{fam} | {pay or '(blank)'} | {cat}"

        # "New Project Needed" — cannot crosswalk; collect for Linear project creation
        if "new project" in v.lower():
            new_projects.append({"payer_family":fam,"insurance_payer":pay,"plan_category":cat,
                                 "note":v,"suggested_project":(r.get("resolved_project") or "").strip()})
            continue

        # UUID in validated = authoritative mapping (name comes from the UUID)
        if UUID_RE.match(v):
            uuid=v; proj=proj_by_uuid.get(uuid)
            if not proj:
                problems.append(f"  UUID not found in projects list ({uuid}): {who}"); continue
        elif v.lower() in TRUTHY:
            proj=(r.get("resolved_project") or "").strip()
            if not proj:
                problems.append(f"  validated 'x' but no project set: {who}"); continue
            uuid=(r.get("project_uuid") or "").strip() or proj_by_name.get(proj,"")
            if not uuid:  # tolerant fallback: exact normalized name match, must be unique
                hits=proj_by_norm.get(pnorm(proj),[])
                if len(hits)==1: proj,uuid=hits[0]
            if not uuid:
                problems.append(f"  project '{proj}' not in projects list: {who}"); continue
        else:
            problems.append(f"  unrecognized validated value '{v}': {who}"); continue

        newrow={"payer_family":fam,"insurance_payer":pay,"plan_category":cat,
                "linear_project":proj,"project_uuid":uuid,"source":"Bruna-validated","effective":month}
        key=(fam,pay,cat)
        if key in xindex:
            ex=xindex[key]
            if ex.get("linear_project","").strip()==proj and ex.get("project_uuid","").strip()==uuid:
                skipped_same+=1; continue
            ex.update(newrow); updated+=1
        else:
            xrows.append(newrow); xindex[key]=newrow; added+=1

    print(f"validated rows applied: +{added} added, ~{updated} updated "
          f"(unchanged {skipped_same}, unmarked {skipped_unmarked})")
    if new_projects:
        print(f'{len(new_projects)} row(s) marked "New Project Needed" — not crosswalked:')
        for n in new_projects: print(f"  · {n['payer_family']} | {n['insurance_payer'] or '(blank)'} | {n['plan_category']}")
    if problems:
        print(f"{len(problems)} validated row(s) could NOT be applied (fix and re-feed):")
        for p in problems[:30]: print(p)

    if a.dry_run:
        print("\n--dry-run: nothing written."); return

    # needs-new-project report (always written when present, even on otherwise-empty runs)
    if new_projects:
        os.makedirs(OUT, exist_ok=True)
        npath=f"{OUT}/New_Projects_Needed_{stamp}.csv"
        with open(npath,"w",newline="") as f:
            w=csv.DictWriter(f,fieldnames=["payer_family","insurance_payer","plan_category","note","suggested_project"])
            w.writeheader(); w.writerows(new_projects)
        print(f"\nwrote needs-new-project report → {npath}")

    if added==0 and updated==0:
        print("crosswalk unchanged."); return
    if os.path.exists(a.crosswalk):
        bak=f"{a.crosswalk}.bak_{stamp}"; shutil.copy2(a.crosswalk,bak)
        print(f"backed up crosswalk → {bak}")
    with open(a.crosswalk,"w",newline="") as f:
        w=csv.DictWriter(f,fieldnames=XCOLS,extrasaction="ignore"); w.writeheader(); w.writerows(xrows)
    print(f"wrote {len(xrows)} crosswalk rows → {a.crosswalk}")
    print("These now win automatically on the next reconcile run.")

if __name__=="__main__":
    main()
