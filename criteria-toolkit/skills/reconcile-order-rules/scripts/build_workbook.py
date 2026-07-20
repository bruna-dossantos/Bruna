#!/usr/bin/env python3
"""Render the reconciliation workbook: Summary + verdict tabs + Needs Review (order names)."""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HDR=PatternFill("solid",fgColor="1F2937"); HF=Font(color="FFFFFF",bold=True)
TITLE=Font(bold=True,size=15); SUB=Font(italic=True,size=9,color="6B7280")
COLORS={"DONE":"D1FAE5","FLIP":"FEF3C7","CONFLICT":"FEE2E2","NEW-TIX":"DBEAFE","UNIT-GAP":"F3E8FF"}
ILLEGAL=re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
def clean(s): return ILLEGAL.sub('',str(s or '')).replace('​','').strip()

def build(detail, res, buckets, vol, out_path, stamp):
    wb=Workbook()
    def hstyle(ws,n):
        for c in range(1,n+1): ws.cell(1,c).fill=HDR; ws.cell(1,c).font=HF
        ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(n)}1"
    def widths(ws,ww):
        for i,w in enumerate(ww,1): ws.column_dimensions[get_column_letter(i)].width=w

    ws=wb.active; ws.title="Summary"
    ws["A1"]="Order-Rule → Linear Reconciliation"; ws["A1"].font=TITLE
    resolved=sum(1 for x in res if x["resolved_project"])
    ws["A2"]=f"Run {stamp} · {resolved:,}/{len(res):,} rows resolved · READ-ONLY (no Linear writes)"; ws["A2"].font=SUB
    legend=[("DONE","Rule built & ticket already Done — in sync"),
            ("FLIP","Rule built, ticket Open → move to Done"),
            ("CONFLICT","Rule built, ticket Canceled/Not-Covered/Blocked → review"),
            ("NEW-TIX","Rule built, no ticket for this code/drug × payer → create"),
            ("UNIT-GAP","Code/drug not in Linear under that title → naming gap")]
    for i,(k,d) in enumerate(legend):
        c=ws.cell(4+i,1,k); c.fill=PatternFill("solid",fgColor=COLORS[k]); c.font=Font(bold=True); ws.cell(4+i,2,d)
    r=4+len(legend)+1
    for i,h in enumerate(["Verdict","DME units","DME rows","Infusion units","Infusion rows"],1):
        c=ws.cell(r,i,h); c.fill=HDR; c.font=HF
    for i,v in enumerate(["DONE","FLIP","CONFLICT","NEW-TIX","UNIT-GAP"]):
        rr=r+1+i; c=ws.cell(rr,1,v); c.fill=PatternFill("solid",fgColor=COLORS[v]); c.font=Font(bold=True)
        ws.cell(rr,2,buckets.get(("DME",v),0)); ws.cell(rr,3,vol.get(("DME",v),0))
        ws.cell(rr,4,buckets.get(("Infusion",v),0)); ws.cell(rr,5,vol.get(("Infusion",v),0))
    widths(ws,[40,14,12,16,14])

    def sheet(name,vs):
        ws=wb.create_sheet(name)
        heads=["Service","Code / Drug","Linear Project","Verdict","Rows built","Basis","Linear state(s)","Ticket ID(s)","URL(s)"]
        ws.append(heads); hstyle(ws,len(heads))
        data=sorted([d for d in detail if d["verdict"] in vs],key=lambda x:(x["service_line"],-x["rows_built"]))
        for d in data:
            ws.append([d["service_line"],clean(d["unit"]),clean(d["project"]),d["verdict"],d["rows_built"],
                       d["basis"],d["linear_states"],d["ticket_ids"],d["ticket_urls"]])
            ws.cell(ws.max_row,4).fill=PatternFill("solid",fgColor=COLORS.get(d["verdict"],"FFFFFF"))
        widths(ws,[9,26,40,10,10,26,24,16,52])
    sheet("FLIP → mark Done",{"FLIP"}); sheet("NEW-TIX → create",{"NEW-TIX"})
    sheet("CONFLICT → review",{"CONFLICT"}); sheet("UNIT-GAP",{"UNIT-GAP"}); sheet("DONE (in sync)",{"DONE"})

    # Needs Review: unresolved rows with order type name + id
    ws=wb.create_sheet("Needs Review (order names)")
    heads=["Reason","Service","Payer Family","Insurance Payer","Category","Plan Type","Code","Service Line Name","ORDER TYPE NAME","Order Type ID","Eligibility ID","→ Correct Payer/Project (fill in)"]
    ws.append(heads); hstyle(ws,len(heads))
    un=[x for x in res if not x["resolved_project"]]
    un.sort(key=lambda x:(x["basis"],x['service_line_category'],x['payer_family'],x.get('service_line','')))
    for x in un:
        ws.append([x["basis"].replace("unresolved:",""),x['service_line_category'],clean(x['payer_family']),
                   clean(x['insurance_payer']),x['plan_category'],clean(x.get('plan_type','')),clean(x['code']),
                   clean(x.get('service_line','')),clean(x['order_rule_name']),clean(x.get('order_rule_id','')),
                   clean(x.get('eligibility_id','')),""])
    widths(ws,[26,10,22,30,16,18,10,24,60,40,16,32])
    wb.save(out_path)
