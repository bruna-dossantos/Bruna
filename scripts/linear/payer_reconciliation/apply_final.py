#!/usr/bin/env python3
"""Apply Bruna's column-J overrides + mine order_rule_name for the 'Need Order Type Info' combos.
Writes an updated row_resolution.csv (family-anchored base + overrides + order-name layer)."""
import csv, re
from collections import Counter, defaultdict
from openpyxl import load_workbook
import family_lib as F

DL="/Users/brunadossantos/Downloads/Service Line Order Type Codes Jul 15 2026.csv"
WORK="/private/tmp/claude-501/-Users-brunadossantos-Projects-Bruna--claude-worktrees-inspiring-cannon-a54635/b4e7b3c8-2754-46fa-a3ad-6e0e1f6f10f1/scratchpad/recon"
XL="/Users/brunadossantos/Documents/Claude/Projects/Criteria Updates/Full_Insurance_Mapping_Jul16.xlsx"

# ---- live projects (superset) ----
projects={r['Name']:r['UUID'] for r in csv.DictReader(open(f"{WORK}/projects_live.csv"))}
uuid2name={v:k for k,v in projects.items()}
def slug(n): return re.sub(r'-+','-',re.sub(r'[^a-z0-9]+','-',n.lower())).strip('-')
slug2name={slug(n):n for n in projects}
UUIDRE=re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$')
def parse_J(j):
    j=(j or '').replace('​','').strip()
    if not j: return (None,None)
    if UUIDRE.match(j): return (uuid2name.get(j,j),j)
    if j.startswith("http"):
        m=re.search(r'/project/(.+?)/',j); s=re.sub(r'-[0-9a-f]{8,}$','',m.group(1)) if m else ""
        if s in slug2name: n=slug2name[s]; return (n,projects[n])
        return (s.replace('-',' ').title(),None)
    if j in projects: return (j,projects[j])
    return (j,None)

# ---- read workbook overrides + need-info combos ----
ws=load_workbook(XL,data_only=True).active
override={}    # (fam,payer,cat) -> (proj_name, uuid)
need_info=set()
for r in range(2,ws.max_row+1):
    d={ws.cell(1,c).value:ws.cell(r,c).value for c in range(1,ws.max_column+1)}
    fam=(d['Payer Family'] or '').strip(); pay=(d['Insurance Payer'] or '').strip(); cat=(d['Category'] or '').strip()
    fam='' if fam=='(blank family)' else fam; pay='' if pay=='(blank payer)' else pay
    ns=(d['Next Steps'] or '').strip()
    if ns=="Update Project Mapping":
        nm,uid=parse_J(d['New Mapping'])
        if nm: override[(fam,pay,cat)]=(nm,uid or projects.get(nm,""))
    elif ns=="Need Order Type Info":
        need_info.add((fam,pay,cat))

# ---- order-name resolver ----
ABBR=F.ABBR; STATES=F.STATES
def states_from(t):
    tl=t.lower(); out=[]
    for s in STATES:
        if re.search(r'\b'+re.escape(s)+r'\b',tl): out.append(s)
    for ab,full in ABBR.items():
        if re.search(r'(?<![a-z])'+ab+r'(?![a-z])',tl) and full not in out: out.append(full)
    return out
BRAND=[  # (keywords in order name) -> project name template with {S}=Title state
 (['fidelis'], "Fidelis Care"),
 (['nebraska total care'], "Nebraska Total Care Medicaid MCO (Nebraska)"),
 (['ambetter'], "Ambetter Centene Medicaid MCO ({S})"),
 (['centene'], "Centene Medicaid MCO ({S})"),
 (['northwood'], "Northwood"),
 (['anthem'], "Anthem Medicaid MCO ({S})"),
 (['bcbs','blue cross'], "Blue Cross Blue Shield Medicaid MCO ({S})"),
 (['uhc','unitedhealthcare','united healthcare','community'], "United Healthcare Community Medicaid MCO ({S})"),
]
def title(s): return " ".join(w.capitalize() for w in s.split())
def resolve_order_name(orn, cat):
    n=(orn or '').lower()
    sts=states_from(orn)
    st=title(sts[0]) if sts else None
    for kws,tmpl in BRAND:
        if any(k in n for k in kws):
            cand=tmpl.replace("{S}",st or "")
            if "{S}" in tmpl and not st: return (None,"ordername:brand-no-state")
            if cand in projects: return (cand,"ordername")
            # fuzzy: match on slug ignoring parenthetical when no exact
            sl=slug(cand)
            if sl in slug2name: return (slug2name[sl],"ordername")
            return (None,f"ordername:no-project:{cand}")
    return (None,"ordername:no-indicator")

# ---- name fallback (from hybrid) ----
def pnorm(s):
    s=(s or '').lower(); s=re.sub(r'\b(of|the|inc|llc|company|health\s?plan)\b','',s)
    s=s.replace("blue cross blue shield","bcbs"); return re.sub(r'[^a-z0-9]+','',s)
nproj=defaultdict(list)
for name in projects: nproj[pnorm(name)].append(name)
mapper={}
for fn in ["dme_mapping.csv","inf_mapping.csv"]:
    for r in csv.DictReader(open(f"{WORK}/{fn}")):
        src=r['source'].strip(); proj=r['project'].strip()
        try: conf=int(r['confidence'])
        except: conf=0
        if src=="Blue Cross Blue Shield of Arkansas": proj="Blue Cross Blue Shield (Arkansas)"; conf=3
        if src=="Blue Cross Blue Shield of Michigan/ Medicare": proj=""; conf=1
        if proj and conf>=2 and proj in projects: mapper.setdefault(src,proj)
def lob_ok(cat,proj):
    if not proj: return True
    pl=proj.lower(); mcd="medicaid" in pl or "mco" in pl; ma=("medicare advantage" in pl) or (re.search(r'\bma\b',pl) is not None); mc=("medicare" in pl) and not ma
    if cat=="COMMERCIAL" and (mcd or ma or mc): return False
    if cat=="MEDICAID" and not mcd: return False
    if cat=="MEDICARE_ADVANTAGE" and not ma: return False
    if cat=="MEDICARE" and not (mc or ma): return False
    return True
def name_resolve(payer,cat):
    payer=(payer or '').strip()
    if not payer: return None
    for cand in ([mapper[payer]] if payer in mapper else [])+([payer] if payer in projects else [])+(nproj[pnorm(payer)] if len(nproj.get(pnorm(payer),[]))==1 else []):
        return cand if lob_ok(cat,cand) else None
    return None

# ---- resolve every row ----
raw=list(csv.DictReader(open(DL)))
out=[]; basis_ct=Counter(); resolved=0
for r in raw:
    fam=r['payer_family'].strip(); pay=r['insurance_payer'].strip(); cat=r['plan_category'].strip()
    orn=r['order_rule_name']
    proj=None; basis=None
    key=(fam,pay,cat)
    if key in override:
        proj,uid=override[key]; basis="override"
    elif key in need_info:
        proj,b=resolve_order_name(orn,cat); basis=b if proj else b
    if not proj and key not in need_info:
        # base family-anchored + name fallback
        p,m=F.resolve(fam,cat,pay,orn)
        if p: proj,basis=p,m
        else:
            nm=name_resolve(pay,cat)
            if nm: proj,basis=nm,"name"
            else: basis="unresolved:"+m
    if not proj and key in need_info and not basis:
        basis="unresolved:need-order-type-info"
    if proj: resolved+=1
    basis_ct[basis]+=1
    out.append([r['service_line_category'],r['code'],r['service_line'],fam,pay,cat,
                proj or "", projects.get(proj,"") if proj else "", basis])

with open(f"{WORK}/row_resolution.csv","w",newline="") as f:
    w=csv.writer(f); w.writerow(["service_line_category","code","service_line","payer_family","insurance_payer","plan_category","resolved_project","project_uuid","basis"])
    w.writerows(out)

n=len(raw)
print(f"total {n:,} | RESOLVED {resolved:,} ({100*resolved/n:.1f}%) | unresolved {n-resolved:,}")
print("\nbasis breakdown:")
for b,v in basis_ct.most_common():
    tag="  " if not str(b).startswith("unresolved") else "x "
    print(f"  {tag}{v:6,}  {b}")
